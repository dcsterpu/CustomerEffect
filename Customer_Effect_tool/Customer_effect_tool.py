from PyQt5.QtWidgets import (QPushButton, QWidget, QLineEdit, QApplication, QLabel, QTextEdit, QFileDialog, QRadioButton)
from PyQt5 import QtCore
import sys
import GeneralStructureTester
import xlrd
from openpyxl import Workbook
import time
from datetime import date
from openpyxl.styles import Alignment, Border, Side
import win32com.client as win32
import win32api
files_path = []


class LineEdit(QLineEdit):

    def __init__(self, title, parent):
        super().__init__(title, parent)

        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [str(u.toLocalFile()) for u in event.mimeData().urls()]
        for f in files:
            temp = self.text()
            temp = temp + str(f) + "\n"
            files_path.append(str(f))
            self.setText(temp)
            # self.setText(("\n"))


class MainWindow(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()



        self.flag_error = False
        self.DOC3WorkBook = None
        self.DOC3Path = ""
        self.DOC4Path = ""
        # self.DOC4WorkBook = ""
        self.dtc_link = dict()
        self.technical_effects = []
        self.technical_dtc = []
        self.DOCs3 = []
        self.tracability = []
        self.customer = []
        self.file_created = False

    def initUI(self):
        # set workspace
        self.tool_version = "Customer Effect List V2.0"
        self.setWindowTitle(self.tool_version)
        self.setGeometry(300, 300, 620, 440)

        # set controls for DOC3
        self.labelDoc3 = QLabel("System TSD File", self)
        self.labelDoc3.move(20, 35)
        self.editDoc3 = QLineEdit('', self)
        self.editDoc3.setDragEnabled(False)
        self.editDoc3.move(120, 30)
        self.editDoc3.resize(430, 25)
        self.buttonDoc3 = QPushButton("Browse", self)
        self.buttonDoc3.move(560, 30)
        self.buttonDoc3.resize(50, 25)
        self.buttonDoc3.clicked.connect(self.openFileNameDialog1)

        # set controls for DOC4
        self.labelDoc4 = QLabel("Function TSD File(s)", self)
        self.labelDoc4.move(20, 100)
        self.editDoc4 = LineEdit('', self)
        self.editDoc4.setDragEnabled(True)
        self.editDoc4.move(120, 70)
        self.editDoc4.resize(490, 80)
        self.editDoc4.setReadOnly(True)

        # controls for output file
        self.labelOutput = QLabel("Output file path", self)
        self.labelOutput.move(20, 165)
        self.editOutput = QLineEdit('', self)
        self.editOutput.setDragEnabled(False)
        self.editOutput.move(120, 160)
        self.editOutput.resize(430, 25)
        self.buttonOutput = QPushButton("Browse", self)
        self.buttonOutput.move(560, 160)
        self.buttonOutput.resize(50, 25)
        self.buttonOutput.clicked.connect(self.openFileNameDialog2)

        #control for file name
        self.labelName = QLabel("Output file name", self)
        self.labelName.move(20, 200)
        self.editName = QLineEdit('', self)
        self.editName.setDragEnabled(False)
        self.editName.move(120, 195)
        self.editName.resize(430, 25)

        self.button = QPushButton("Generate", self)
        self.button.move(220, 400)
        self.button.clicked.connect(self.buttonGenerateClicked)

        self.button2 = QPushButton("Open file", self)
        self.button2.move(320, 400)
        self.button2.setEnabled(False)
        self.button2.clicked.connect(self.buttonOpenFileClicked)

        self.labelTypeGeneration = QLabel("File generation strategy:", self)
        self.labelTypeGeneration.move(20, 235)
        self.RadioButtonTehnicalEffects = QRadioButton(self)
        self.RadioButtonTehnicalEffects.setText("Technical effects based")
        self.RadioButtonTehnicalEffects.setChecked(True)
        self.RadioButtonDTC = QRadioButton(self)
        self.RadioButtonDTC.setText("DTC based")
        self.RadioButtonTehnicalEffects.move(165, 235)
        self.RadioButtonDTC.move(165, 260)

        message = ""
        self.textbox = QTextEdit(self)
        self.textbox.setText(message)
        self.textbox.move(10, 285)
        self.textbox.resize(600, 100)
        self.textbox.setReadOnly(True)

    def openFileNameDialog1(self):
        fileName1, _filter = QFileDialog.getOpenFileName(self, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.editDoc3.setText(fileName1)

    def openFileNameDialog2(self):
        fileName2 = QFileDialog.getExistingDirectory(self, 'Save File', QtCore.QDir.rootPath(), QFileDialog.ShowDirsOnly)
        self.editOutput.setText(fileName2)

    def DOC4Parser(self):

        self.DOC4List = []

        self.DOC4Path = self.editDoc3.text()
        filename = self.DOC4Path.split("/")[-1]
        text = self.textbox.toPlainText()
        if text == "":
            self.textbox.setText(filename)
        else:
            self.textbox.setText(text + "\n" + filename)

        if self.DOC4Path.split(".")[-1] == "xls":
            self.DOC4WorkBook = xlrd.open_workbook(self.DOC4Path, formatting_info=True)
        else:
            self.DOC4WorkBook = xlrd.open_workbook(self.DOC4Path)

        sheets = self.DOC4WorkBook.sheet_names()
        if "Technical effect" in sheets or "Effets techniques" in sheets:
            text = self.textbox.toPlainText()
            self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0150 " + " OK")
            try:
                workSheet = self.DOC4WorkBook.sheet_by_name("Technical effect")
            except:
                workSheet = self.DOC4WorkBook.sheet_by_name("Effets techniques")

            nrCols = workSheet.ncols
            nrRows = workSheet.nrows
            refNameCol = -1
            refNameRow = -1
            refReqCol = -1
            refReqRow = -1
            refAccRow = -1
            refAccCol = -1

            for index1 in range(0,nrRows):
                for index2 in range(0,nrCols):
                    if str(workSheet.cell(index1, index2).value).casefold().strip() == "Noms".casefold() or str(workSheet.cell(index1, index2).value).casefold().strip() == "Names".casefold() or str(workSheet.cell(index1, index2).value).casefold().strip() == "Name".casefold():
                        refNameRow = index1
                        refNameCol = index2
                    if str(workSheet.cell(index1, index2).value).casefold().strip() == "Upstream requirements".casefold() or str(workSheet.cell(index1, index2).value).casefold().strip() == "Référence amont".casefold():
                        refReqRow = index1
                        refReqCol = index2
                    if str(workSheet.cell(index1, index2).value).casefold().strip() == "Taken into account".casefold():
                        refAccRow = index1
                        refAccCol = index2

            if refNameRow != -1 and refNameCol != -1:
                text = self.textbox.toPlainText()
                self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0160 " + " OK")
            else:
                text = self.textbox.toPlainText()
                self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0160 " + " NOK")

            if refReqRow != -1 and refReqCol != -1:
                text = self.textbox.toPlainText()
                self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0170 " + " OK")
            else:
                text = self.textbox.toPlainText()
                self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0170 " + " NOK")

            if refNameRow != -1 and refNameCol != -1 and refReqRow != -1 and refReqCol != -1 and refAccRow != -1 and refAccCol != -1:
                for index in range(refNameRow + 1, nrRows):
                    if workSheet.cell(index, refNameCol).value != "" and ( str(workSheet.cell(index, refAccCol).value).casefold().strip() == "oui" or str(workSheet.cell(index, refAccCol).value).casefold().strip() == "yes" ):
                        dict = {}
                        dict["technical"] = workSheet.cell(index, refNameCol).value
                        dict["upstream"] = workSheet.cell(index, refReqCol).value
                        self.technical_effects.append(dict)

        else:
            text = self.textbox.toPlainText()
            self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0150 " + " NOK")

        if "Table" in sheets or "Tableau" in sheets:

            try:
                workSheet = self.DOC4WorkBook.sheet_by_name("Table")
            except:
                workSheet = self.DOC4WorkBook.sheet_by_name("Tableau")

            nrCols = workSheet.ncols
            nrRows = workSheet.nrows
            refDTCCol = -1
            refDTCRow = -1
            refTehnicalCol = -1
            refTehnicalRow = -1
            refLinkCol = -1
            refLinkRow = -1


            for index1 in range(0,nrRows):
                for index2 in range(0,nrCols):
                    if str(workSheet.cell(index1, index2).value).casefold().strip() == "Data Trouble code".casefold():
                        refDTCRow = index1
                        refDTCCol = index2
                    if str(workSheet.cell(index1, index2).value).casefold().strip() == "Technical effect".casefold():
                        refTehnicalRow = index1
                        refTehnicalCol = index2
                    if str(workSheet.cell(index1, index2).value).casefold().strip() == "Link to another DST".casefold():
                        refLinkRow = index1
                        refLinkCol = index2

            if refDTCRow != -1 and refDTCCol != -1:
                text = self.textbox.toPlainText()
                self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0230 " + " OK")
            else:
                text = self.textbox.toPlainText()
                self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0230 " + " NOK")


            if refDTCRow != -1 and refDTCCol != -1 and refTehnicalRow != -1 and refTehnicalCol != -1:
                for index in range(refTehnicalRow + 1, nrRows):
                    cel = workSheet.cell(index, refTehnicalCol).value
                    if cel != "":
                        if ";" in cel or "," in cel or "\n" in cel:
                            # cel = cel.replace(",", "<>").replace(";", "<>").replace("\n", "<>")
                            cel = cel.replace(";", "<>").replace("\n", " ")
                            cel = cel.split("<>")
                            for elem in cel:
                                dict = {}
                                dict['technical'] = elem.strip()
                                dict['dtc'] = workSheet.cell(index, refDTCCol).value
                                self.technical_dtc.append(dict)

                        else:
                            dict = {}
                            dict['technical'] = workSheet.cell(index, refTehnicalCol).value.replace("\n", " ")
                            dict['dtc'] = workSheet.cell(index, refDTCCol).value
                            self.technical_dtc.append(dict)
            else:
                # text = self.textbox.toPlainText()
                self.textbox.setText("The file could not be generated because the connection cannot be made between 'Technical effect' and 'Data Trouble code'")
                win32api.MessageBox(0, 'File generation failed!', 'Error')
                self.flag_error = True

            if self.RadioButtonDTC.isChecked() == True:
                if refDTCRow != -1 and refDTCCol != -1 and refLinkCol != -1 and refLinkRow != -1:
                    for index in range(refDTCRow + 1, nrRows):
                        cel = workSheet.cell(index, refDTCCol).value
                        if cel != "" and cel not in self.dtc_link:
                            inter_list = []
                            for index_inter in range(index, nrRows):
                                if workSheet.cell(index_inter, refDTCCol).value == cel:
                                    if workSheet.cell(index_inter, refLinkCol).value != "" and workSheet.cell(index_inter, refLinkCol).value not in inter_list and "\n" not in workSheet.cell(index_inter, refLinkCol).value:
                                        inter_list.append(workSheet.cell(index_inter, refLinkCol).value)
                                    elif workSheet.cell(index_inter, refLinkCol).value != "" and workSheet.cell(index_inter, refLinkCol).value not in inter_list and "\n" in workSheet.cell(index_inter, refLinkCol).value:
                                        reqs = workSheet.cell(index_inter, refLinkCol).value.split("\n")
                                        for req in reqs:
                                            if req not in inter_list:
                                                inter_list.append(req)
                            # if inter_list:
                            self.dtc_link[cel] = inter_list

                else:
                    self.textbox.setText("The file could not be generated because the connection cannot be made between 'Data Trouble code' and 'Link to another DST")
                    win32api.MessageBox(0, 'File generation failed!', 'Error')
                    self.flag_error = True

        else:
            text = self.textbox.toPlainText()
            self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0230 " + " NOK")

    def DOC3Parser(self):

        self.DOC3List = []

        self.DOC3Path = self.editDoc4.text().split("\n")
        for DOC3 in self.DOC3Path:
            if DOC3 != "":
                filename = DOC3.split("/")[-1]
                text = self.textbox.toPlainText()
                if text == "":
                    self.textbox.setText(filename)
                else:
                    self.textbox.setText(text + "\n" + filename)

                if DOC3.split(".")[-1] == "xls":
                    DOC3WorkBook = xlrd.open_workbook(DOC3, formatting_info=True)
                else:
                    DOC3WorkBook = xlrd.open_workbook(DOC3)

                sheets = DOC3WorkBook.sheet_names()
                if "Req. of tech. effects" in sheets or "Effets techniques" in sheets:
                    text = self.textbox.toPlainText()
                    self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0200 " + " OK")
                    try:
                        workSheet = DOC3WorkBook.sheet_by_name("Req. of tech. effects")
                    except:
                        workSheet = DOC3WorkBook.sheet_by_name("Effets techniques")

                    nrCols = workSheet.ncols
                    nrRows = workSheet.nrows
                    refNameCol = -1
                    refNameRow = -1
                    refTracCol = -1
                    refTracRow = -1

                    for index1 in range(0, nrRows):
                        for index2 in range(0, nrCols):
                            if str(workSheet.cell(index1, index2).value).casefold().strip() == "Reference".casefold() or str(workSheet.cell(index1, index2).value).casefold().strip() == "Names".casefold():
                                refNameRow = index1
                                refNameCol = index2
                            if str(workSheet.cell(index1, index2).value).casefold().strip() == "Tracability with the TSD".casefold() or str(workSheet.cell(index1, index2).value).casefold().strip() == "Référence amont".casefold():
                                refTracRow = index1
                                refTracCol = index2


                    if refNameRow != -1 and refNameCol != -1:
                        text = self.textbox.toPlainText()
                        self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0210 " + " OK")
                    else:
                        text = self.textbox.toPlainText()
                        self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0210 " + " NOK")

                    if refTracRow != -1 and refTracCol != -1:
                        text = self.textbox.toPlainText()
                        self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0220 " + " OK")
                    else:
                        text = self.textbox.toPlainText()
                        self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0220 " + " NOK")

                    if refNameCol != -1 and refTracCol != -1:
                        for index in range(refNameRow + 1, nrRows):
                            dict1 = {}
                            if ";" in workSheet.cell(index, refTracCol).value:
                                values = workSheet.cell(index, refTracCol).value.split(";")
                                values_good = []
                                for elem in values:
                                    values_good.append(elem.split("(")[0])
                                dict1[workSheet.cell(index, refNameCol).value] = values_good
                                self.tracability.append(dict1)
                            else:
                                dict1[workSheet.cell(index, refNameCol).value] = workSheet.cell(index, refTracCol).value.split("(")[0]
                                self.tracability.append(dict1)


                    if "Table" in sheets:
                        workSheetTable = DOC3WorkBook.sheet_by_name("Table")
                        tableRefCol = -1
                        tableRefRow = -1
                        tableCustomerCol = -1
                        tableCustomerRow = -1

                        for index1 in range(0, workSheetTable.nrows):
                            for index2 in range(0, workSheetTable.ncols):
                                if str(workSheetTable.cell(index1, index2).value).casefold().strip() == "Reference".casefold():
                                    tableRefRow = index1
                                    tableRefCol = index2
                                if str(workSheetTable.cell(index1, index2).value).casefold().strip() == "Customer effect".casefold():
                                    tableCustomerRow = index1
                                    tableCustomerCol = index2
                                if tableRefCol != -1 and tableCustomerCol != -1:
                                    break
                            if tableRefCol != -1 and tableCustomerCol != -1:
                                break

                    if tableRefCol != -1 and tableCustomerCol != -1:
                        for index in range(tableRefRow + 1, workSheetTable.nrows):
                            dict2 = {}
                            dict2["tracability"] = workSheetTable.cell(index, tableRefCol).value
                            dict2["customer"] = workSheetTable.cell(index, tableCustomerCol).value
                            self.customer.append(dict2)


                else:
                    text = self.textbox.toPlainText()
                    self.textbox.setText(text + "\n" + "     Test_02043_19_04175_STRUCT_0200 " + " NOK")

    def CreateFileTechnicalEffectsBased(self):


        wb = Workbook()

        ws1 = wb.create_sheet("Customer Effects List")
        ws2 = wb.create_sheet("Report information")

        ws2['A1'] = "Tool version:"
        ws2['B1'] = self.tool_version

        ws2['A3'] = "Date of the test:"
        ws2['B3'] = self.start_date.strftime("%d/%m/%Y")

        ws2['A4'] = "Time of the test:"
        ws2['B4'] = time.strftime('%H:%M:%S', self.start_time)

        ws2['A7'] = "TSD system file used:"
        ws2['B7'] = self.DOC4Path.split("/")[-1]

        ws2['A8'] = "TSD function file used:"
        ws2['B8'] = self.DOC3Path[0].split("/")[-1]
        if len(self.DOC3Path) > 2:
            for i in range(1,len(self.DOC3Path)):
                index = 'B' + str(8 + i)
                if self.DOC3Path[i] != "":
                    ws2[index] = self.DOC3Path[i].split("/")[-1]



        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 100

        if "Sheet" in wb.sheetnames:
            ws = wb["Sheet"]
            wb.remove(ws)

        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 35
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['E'].width = 25
        ws1.column_dimensions['F'].width = 25
        ws1.column_dimensions['G'].width = 90
        ws1.column_dimensions['H'].width = 90



        # for index1 in range(1, len(self.customer)+1):
        for index1 in range(1, 3000):
            for index2 in range(1, 9):
                ws1.cell(index1, index2).border = Border(top=Side(border_style='thin', color='00000000'), right=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'),left=Side(border_style='thin', color='00000000'))

        ws1.merge_cells('A1:D1')
        ws1['A1'] = "System TSD"
        ws1['A1'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('E1:H1')
        ws1['E1'] = "Function TSD"
        ws1['E1'].alignment = Alignment(horizontal='center')

        ws1['A2'] = "Data Trouble Code"
        ws1['A2'].alignment = Alignment(horizontal='center')

        ws1['B2'] = "Technical effects"
        ws1['B2'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('C2:D2')
        ws1['C2'] = "Upstream requirements"
        ws1['C2'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('E2:F2')
        ws1['E2'] = "Tracability with the TSD"
        ws1['E2'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('G2:H2')
        ws1['G2'] = "Customer effects"
        ws1['G2'].alignment = Alignment(horizontal='center')


        if self.technical_effects:
            current_row = 3
            for elem in self.technical_effects:

                initial_row = current_row
                list_dtc = []
                for dtcs in self.technical_dtc:
                    if dtcs['technical'].strip() == elem['technical'].replace("\n", " ").strip():
                        if (dtcs['dtc'] + "\n") not in list_dtc:
                            list_dtc.append(dtcs['dtc'] + "\n")

                dtc_value = ""
                for dtc_elem in list_dtc:
                    dtc_value = dtc_value + dtc_elem

                index1 = 'A' + str(current_row)
                ws1[index1] = dtc_value
                ws1[index1].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)


                index2 = 'B' + str(current_row)
                ws1[index2] = elem["technical"]
                ws1[index2].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

                index3 = 'C' + str(current_row)
                ws1[index3] = elem["upstream"]
                ws1[index3].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

                tracability_list = []
                if elem["upstream"] == "":
                    index4 = 'D' + str(current_row)
                    ws1[index4] = ""
                    current_row = current_row + 1
                    ws1[index4].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                else:
                    upstreams = elem["upstream"].split("\n")
                    for stream in upstreams:

                        intermediate_row = current_row

                        index4 = 'D' + str(current_row)
                        if stream != "":
                            ws1[index4] = stream
                            ws1[index4].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)


                        for elem in self.tracability:
                            try:
                                tracability = elem[ws1.cell(current_row, 4).value]
                                if isinstance(tracability, str):
                                    index5 = 'E' + str(intermediate_row)
                                    ws1[index5] = tracability
                                    ws1[index5].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                                    intermediate_row += 1
                                    if (tracability + ";\n") not in tracability_list:
                                        tracability_list.append(tracability + ";\n")
                                    break
                                else:
                                    for trac_elem in tracability:
                                        index5 = 'E' + str(intermediate_row)
                                        ws1[index5] = trac_elem
                                        ws1[index5].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                                        intermediate_row += 1
                                        if (trac_elem + ";\n") not in tracability_list:
                                            tracability_list.append(trac_elem + ";\n")
                                    break
                            except:
                                pass

                        if intermediate_row - current_row > 1:
                            ws1.merge_cells('D' + str(current_row) + ":D" + str(intermediate_row - 1))
                        elif intermediate_row - current_row == 0 and stream != "":
                            intermediate_row += 1

                        current_row = intermediate_row

                value = ""
                index6 = "F" + str(initial_row)
                for trac in tracability_list:
                    value = value + trac
                ws1[index6] = value
                ws1[index6].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

                customer_list = []
                for index in range(initial_row, current_row):
                    for elem in self.customer:
                        if ws1.cell(index, 5).value == elem["tracability"]:
                            index7 = 'G' + str(index)
                            ws1[index7] = elem["customer"]
                            ws1[index7].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                            if elem["customer"] + "\n" not in customer_list:
                                customer_list.append(elem["customer"] + "\n")
                            break

                cel_value = ""
                index8 = "H" + str(initial_row)
                for elem in customer_list:
                    cel_value = cel_value + elem
                ws1[index8] = cel_value
                ws1[index8].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                try:
                    ws1.merge_cells('H' + str(initial_row) + ":H" + str(current_row - 1))
                    ws1.merge_cells('F' + str(initial_row) + ":F" + str(current_row - 1))

                    ws1.merge_cells('A' + str(initial_row) + ":A" + str(current_row - 1))
                    ws1.merge_cells('B' + str(initial_row) + ":B" + str(current_row - 1))
                    ws1.merge_cells('C' + str(initial_row) + ":C" + str(current_row - 1))
                except:
                    pass



        self.path_to_save = self.editOutput.text() + "/" + self.editName.text() + ".xlsx"
        wb.save(self.path_to_save)

        self.file_created = True
        if self.file_created:
            win32api.MessageBox(0, 'File has been created!', 'Information')
            self.button2.setEnabled(True)

    def CreateFileDTCBased(self):
        print("Testing...")

        wb = Workbook()

        ws1 = wb.create_sheet("Customer Effects List")
        ws2 = wb.create_sheet("Report information")

        ws2['A1'] = "Tool version:"
        ws2['B1'] = self.tool_version

        ws2['A3'] = "Date of the test:"
        ws2['B3'] = self.start_date.strftime("%d/%m/%Y")

        ws2['A4'] = "Time of the test:"
        ws2['B4'] = time.strftime('%H:%M:%S', self.start_time)

        ws2['A7'] = "TSD system file used:"
        ws2['B7'] = self.DOC4Path.split("/")[-1]

        ws2['A8'] = "TSD function file used:"
        ws2['B8'] = self.DOC3Path[0].split("/")[-1]
        if len(self.DOC3Path) > 2:
            for i in range(1, len(self.DOC3Path)):
                index = 'B' + str(8 + i)
                if self.DOC3Path[i] != "":
                    ws2[index] = self.DOC3Path[i].split("/")[-1]

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 100

        if "Sheet" in wb.sheetnames:
            ws = wb["Sheet"]
            wb.remove(ws)

        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 35
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['E'].width = 25
        ws1.column_dimensions['F'].width = 25
        ws1.column_dimensions['G'].width = 90
        ws1.column_dimensions['H'].width = 90

        for index1 in range(1, 3000):
            for index2 in range(1, 9):
                ws1.cell(index1, index2).border = Border(top=Side(border_style='thin', color='00000000'), right=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'),left=Side(border_style='thin', color='00000000'))

        ws1.merge_cells('A1:D1')
        ws1['A1'] = "System TSD"
        ws1['A1'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('E1:H1')
        ws1['E1'] = "Function TSD"
        ws1['E1'].alignment = Alignment(horizontal='center')

        ws1['A2'] = "Data Trouble Code"
        ws1['A2'].alignment = Alignment(horizontal='center')

        ws1['B2'] = "Technical effects"
        ws1['B2'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('C2:D2')
        ws1['C2'] = "Upstream requirements"
        ws1['C2'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('E2:F2')
        ws1['E2'] = "Tracability with the TSD"
        ws1['E2'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('G2:H2')
        ws1['G2'] = "Customer effects"
        ws1['G2'].alignment = Alignment(horizontal='center')

        if self.dtc_link:
            current_row = 3

            for elem in self.dtc_link:
                initial_row = current_row

                index1 = 'A' + str(current_row)
                ws1[index1] = elem
                ws1[index1].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

                # index2 = 'B' + str(current_row)
                # ws1[index2] = elem["technical"]
                # ws1[index2].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

                index3 = 'C' + str(current_row)
                upstream_req_string = ""
                for i in range(0, len(self.dtc_link[elem])):
                    if upstream_req_string == "":
                        upstream_req_string = self.dtc_link[elem][i]
                    else:
                        upstream_req_string += "\n" + self.dtc_link[elem][i]
                ws1[index3] = upstream_req_string
                ws1[index3].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)


                tracability_list = []
                if upstream_req_string == "":
                    index4 = 'D' + str(current_row)
                    ws1[index4] = ""
                    current_row = current_row + 1
                    ws1[index4].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                else:
                    upstreams = upstream_req_string.split("\n")
                    for stream in upstreams:

                        intermediate_row = current_row

                        index4 = 'D' + str(current_row)
                        if stream != "":
                            ws1[index4] = stream
                            ws1[index4].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

                        for elem in self.tracability:
                            try:
                                tracability = elem[ws1.cell(current_row, 4).value]
                                if isinstance(tracability, str):
                                    index5 = 'E' + str(intermediate_row)
                                    ws1[index5] = tracability
                                    ws1[index5].alignment = Alignment(vertical='center', horizontal='center',
                                                                      wrap_text=True)
                                    intermediate_row += 1
                                    if (tracability + ";\n") not in tracability_list:
                                        tracability_list.append(tracability + ";\n")
                                    break
                                else:
                                    for trac_elem in tracability:
                                        index5 = 'E' + str(intermediate_row)
                                        ws1[index5] = trac_elem
                                        ws1[index5].alignment = Alignment(vertical='center', horizontal='center',
                                                                          wrap_text=True)
                                        intermediate_row += 1
                                        if (trac_elem + ";\n") not in tracability_list:
                                            tracability_list.append(trac_elem + ";\n")
                                    break
                            except:
                                pass

                        if intermediate_row - current_row > 1:
                            ws1.merge_cells('D' + str(current_row) + ":D" + str(intermediate_row - 1))
                        elif intermediate_row - current_row == 0 and stream != "":
                            intermediate_row += 1

                        current_row = intermediate_row

                value = ""
                index6 = "F" + str(initial_row)
                for trac in tracability_list:
                    value = value + trac
                ws1[index6] = value
                ws1[index6].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

                customer_list = []
                for index in range(initial_row, current_row):
                    for elem in self.customer:
                        if ws1.cell(index, 5).value == elem["tracability"]:
                            index7 = 'G' + str(index)
                            ws1[index7] = elem["customer"]
                            ws1[index7].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                            if elem["customer"] + "\n" not in customer_list:
                                customer_list.append(elem["customer"] + "\n")
                            break

                cel_value = ""
                index8 = "H" + str(initial_row)
                for elem in customer_list:
                    cel_value = cel_value + elem
                ws1[index8] = cel_value
                ws1[index8].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                try:
                    ws1.merge_cells('H' + str(initial_row) + ":H" + str(current_row - 1))
                    ws1.merge_cells('F' + str(initial_row) + ":F" + str(current_row - 1))

                    ws1.merge_cells('A' + str(initial_row) + ":A" + str(current_row - 1))
                    ws1.merge_cells('B' + str(initial_row) + ":B" + str(current_row - 1))
                    ws1.merge_cells('C' + str(initial_row) + ":C" + str(current_row - 1))
                except:
                    pass




        self.path_to_save = self.editOutput.text() + "/" + self.editName.text() + ".xlsx"
        wb.save(self.path_to_save)

        self.file_created = True
        if self.file_created:
            win32api.MessageBox(0, 'File has been created!', 'Information')
            self.button2.setEnabled(True)

    def buttonGenerateClicked(self):

        self.textbox.setText("")
        self.button2.setEnabled(False)
        QApplication.processEvents()
        self.start_time = time.localtime()
        self.start_date = date.today()

        self.DOC4Parser()
        if self.flag_error:
            return
        self.DOC3Parser()

        if self.RadioButtonTehnicalEffects.isChecked() == True:
            self.CreateFileTechnicalEffectsBased()

        elif self.RadioButtonDTC.isChecked() == True:
            self.CreateFileDTCBased()

    def buttonOpenFileClicked(self):

        self.excel = win32.gencache.EnsureDispatch('Excel.Application')

        if self.file_created:

            self.excel.Visible = True
            self.excel.Workbooks.Open(self.path_to_save)



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainWindow()
    ex.show()
    app.exec_()
